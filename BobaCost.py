import openpyxl as xl

workbk = xl.load_workbook(r'E:\PycharmProjects\HeyWorld\Boba.xlsx')
sheet = workbk['Sheet1']

for row in range(3, sheet.max_row + 1):
    cell_price = sheet.cell(row, 5) #col 5 for price
    cell_no_of_visits = sheet.cell(row, 6)
    corrected_price = cell_price.value * cell_no_of_visits.value
    corrected_price_cell = sheet.cell(row, 7)  # col 7
    corrected_price_cell.value = corrected_price


workbk.save(r'E:\PycharmProjects\HeyWorld\BobaUpdated.xlsx')


